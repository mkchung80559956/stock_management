"""
Watchlist Pro v2.0  —  台股自選股專業管理系統
Professional Taiwan stock watchlist management tool.
"""
# ══════════════════════════════════════════════
# IMPORTS
# ══════════════════════════════════════════════
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import yfinance as yf
import json, os, pytz, _io
import urllib.request as _req
import urllib.parse as _parse
import re as _re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import io

# ══════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════
st.set_page_config(
    page_title="Watchlist Pro",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_VERSION = "2.0"
APP_UPDATED = "2026-04-15"
TZ = pytz.timezone("Asia/Taipei")

def now_tw(): return datetime.now(TZ)
def today_str(): return now_tw().strftime("%Y-%m-%d")

# Persistence paths
_WL_FILE      = "/tmp/wlpro_watchlist.json"
_NOTES_FILE   = "/tmp/wlpro_notes.json"
_PORT_FILE    = "/tmp/wlpro_portfolio.json"
_ALERT_FILE   = "/tmp/wlpro_alerts_sent.json"

# ══════════════════════════════════════════════
# CSS — Professional Dark Trading Terminal
# ══════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;600;700&display=swap');

/* ═══ TOKENS ═══ */
:root {
  --bg0:     #04070f;
  --bg1:     #080d18;
  --bg2:     #0c1220;
  --bg3:     #101928;
  --bg4:     #162035;
  --glass:   rgba(16,25,40,0.85);
  --line:    rgba(0,200,255,0.12);
  --line2:   rgba(0,200,255,0.06);
  --text:    #dce8f5;
  --muted:   #4d6880;
  --dim:     #253545;
  --accent:  #00c8ff;
  --green:   #00f5a0;
  --red:     #ff3b52;
  --gold:    #ffc94d;
  --orange:  #ff8c42;
  --purple:  #a78bfa;
  --mono:    'JetBrains Mono', monospace;
  --sans:    'Space Grotesk', sans-serif;
  --r4: 4px; --r8: 8px; --r12: 12px;
}

/* ═══ RESET ═══ */
html, body, .stApp, [data-testid="stAppViewContainer"],
[data-testid="stMain"] > div { background: var(--bg0) !important; }
* { font-family: var(--sans) !important; box-sizing: border-box; }
code, .mono { font-family: var(--mono) !important; }

/* ═══ SCROLLBAR ═══ */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg1); }
::-webkit-scrollbar-thumb { background: var(--dim); border-radius: 2px; }

/* ═══ SIDEBAR ═══ */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, var(--bg1) 0%, var(--bg0) 100%) !important;
  border-right: 1px solid var(--line) !important;
  backdrop-filter: blur(12px);
}
section[data-testid="stSidebar"] * { color: var(--text) !important; }
section[data-testid="stSidebar"] > div { padding-top: 1rem !important; }

/* ═══ TABS ═══ */
.stTabs [data-baseweb="tab-list"] {
  background: transparent !important;
  border-bottom: 1px solid var(--line) !important;
  gap: 0; padding: 0;
}
.stTabs [data-baseweb="tab"] {
  font-family: var(--mono) !important;
  font-size: 0.72rem !important;
  font-weight: 600 !important;
  letter-spacing: 0.1em;
  color: var(--muted) !important;
  padding: 10px 16px !important;
  border: none !important;
  border-bottom: 2px solid transparent !important;
  background: transparent !important;
  text-transform: uppercase;
  transition: color 0.2s, border-color 0.2s;
}
.stTabs [aria-selected="true"] {
  color: var(--accent) !important;
  border-bottom: 2px solid var(--accent) !important;
}
.stTabs [data-baseweb="tab-panel"] { padding-top: 20px !important; }

/* ═══ METRICS ═══ */
div[data-testid="stMetric"] {
  background: var(--bg2) !important;
  border: 1px solid var(--line) !important;
  border-radius: var(--r8) !important;
  padding: 14px 16px !important;
  position: relative; overflow: hidden;
}
div[data-testid="stMetric"]::before {
  content: '';
  position: absolute; top: 0; left: 0;
  width: 3px; height: 100%;
  background: linear-gradient(180deg, var(--accent), transparent);
}
div[data-testid="stMetricLabel"] {
  font-family: var(--mono) !important;
  font-size: 0.65rem !important;
  color: var(--muted) !important;
  letter-spacing: 0.12em;
  text-transform: uppercase;
}
div[data-testid="stMetricValue"] {
  font-family: var(--mono) !important;
  font-size: 1.25rem !important;
  font-weight: 700 !important;
  color: var(--text) !important;
}

/* ═══ DATAFRAME ═══ */
.stDataFrame iframe { border: none !important; }
.stDataFrame [data-testid="stDataFrameResizable"] {
  border: 1px solid var(--line) !important;
  border-radius: var(--r8) !important;
  overflow: hidden;
}

/* ═══ BUTTONS ═══ */
.stButton > button {
  font-family: var(--mono) !important;
  font-size: 0.78rem !important;
  font-weight: 600 !important;
  letter-spacing: 0.06em !important;
  text-transform: uppercase !important;
  background: var(--bg3) !important;
  border: 1px solid var(--dim) !important;
  color: var(--muted) !important;
  border-radius: var(--r8) !important;
  min-height: 44px !important;
  transition: all 0.15s ease !important;
}
.stButton > button:hover {
  border-color: var(--accent) !important;
  color: var(--accent) !important;
  background: rgba(0,200,255,0.06) !important;
  box-shadow: 0 0 16px rgba(0,200,255,0.12) !important;
}
.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, rgba(0,200,255,0.15), rgba(0,200,255,0.08)) !important;
  border-color: var(--accent) !important;
  color: var(--accent) !important;
  box-shadow: 0 0 20px rgba(0,200,255,0.1) !important;
  min-height: 50px !important;
  font-size: 0.85rem !important;
}
.stButton > button[kind="primary"]:hover {
  background: linear-gradient(135deg, rgba(0,200,255,0.25), rgba(0,200,255,0.12)) !important;
  box-shadow: 0 0 30px rgba(0,200,255,0.2) !important;
}

/* ═══ INPUTS ═══ */
.stTextInput > label,
.stSelectbox > label,
.stNumberInput > label,
.stTextArea > label { color: var(--muted) !important; font-size: 0.72rem !important; letter-spacing: 0.06em; }
.stTextInput input, .stNumberInput input {
  background: var(--bg3) !important;
  border: 1px solid var(--dim) !important;
  border-radius: var(--r8) !important;
  color: var(--text) !important;
  font-family: var(--mono) !important;
  font-size: 0.95rem !important;
  min-height: 46px !important;
  padding: 0 14px !important;
  transition: border-color 0.2s, box-shadow 0.2s;
}
.stTextInput input:focus, .stNumberInput input:focus {
  border-color: var(--accent) !important;
  box-shadow: 0 0 0 3px rgba(0,200,255,0.12) !important;
  outline: none !important;
}
.stTextArea textarea {
  background: var(--bg3) !important;
  border: 1px solid var(--dim) !important;
  border-radius: var(--r8) !important;
  color: var(--text) !important;
  font-size: 0.88rem !important;
}
[data-baseweb="select"] > div {
  background: var(--bg3) !important;
  border: 1px solid var(--dim) !important;
  border-radius: var(--r8) !important;
  color: var(--text) !important;
  min-height: 46px !important;
}
[data-baseweb="select"] > div:focus-within {
  border-color: var(--accent) !important;
  box-shadow: 0 0 0 3px rgba(0,200,255,0.1) !important;
}
[data-baseweb="popover"] { background: var(--bg3) !important; border: 1px solid var(--line) !important; }
[role="option"] { background: var(--bg3) !important; color: var(--text) !important; }
[role="option"]:hover { background: var(--bg4) !important; }

/* ═══ EXPANDERS ═══ */
.streamlit-expanderHeader {
  background: var(--bg2) !important;
  border: 1px solid var(--line) !important;
  border-radius: var(--r8) !important;
  color: var(--muted) !important;
  font-size: 0.8rem !important;
  font-weight: 600 !important;
  letter-spacing: 0.04em;
}
.streamlit-expanderContent {
  background: var(--bg1) !important;
  border: 1px solid var(--line);
  border-top: none;
  border-radius: 0 0 var(--r8) var(--r8) !important;
}

/* ═══ ALERTS ═══ */
.stSuccess { background: rgba(0,245,160,0.08) !important; border-left: 3px solid var(--green) !important; border-radius: var(--r8) !important; }
.stInfo    { background: rgba(0,200,255,0.08) !important; border-left: 3px solid var(--accent) !important; border-radius: var(--r8) !important; }
.stWarning { background: rgba(255,140,66,0.08) !important; border-left: 3px solid var(--orange) !important; border-radius: var(--r8) !important; }
.stError   { background: rgba(255,59,82,0.08) !important; border-left: 3px solid var(--red) !important; border-radius: var(--r8) !important; }

/* ═══ TOGGLE / CHECKBOX ═══ */
[data-testid="stCheckbox"] label { color: var(--text) !important; }

/* ═══ DIVIDER ═══ */
hr { border-color: var(--line) !important; }

/* ═══ HIDE CHROME ═══ */
#MainMenu, footer { display: none !important; }
[data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"] { display: none !important; }

/* ═══ LAYOUT ═══ */
.block-container {
  padding: 1rem 1.2rem 2rem 1.2rem !important;
  max-width: 1200px !important;
}

/* ═══ MOBILE ═══ */
@media (max-width: 640px) {
  .block-container { padding: 0.75rem 0.6rem 1.5rem 0.6rem !important; }
  .stTabs [data-baseweb="tab"] { padding: 8px 10px !important; font-size: 0.65rem !important; }
  div[data-testid="stMetricValue"] { font-size: 1.05rem !important; }
}

/* ═══ CUSTOM COMPONENTS ═══ */
.price-card {
  background: var(--bg2);
  border: 1px solid var(--line);
  border-radius: var(--r12);
  padding: 14px 16px;
  margin-bottom: 12px;
  position: relative;
  overflow: hidden;
  transition: transform 0.15s, border-color 0.15s, box-shadow 0.15s;
}
.price-card::after {
  content: '';
  position: absolute; top: 0; right: 0;
  width: 40%; height: 100%;
  background: linear-gradient(90deg, transparent, rgba(0,200,255,0.02));
  pointer-events: none;
}
.price-card.up   { border-left: 3px solid var(--red);    background: linear-gradient(135deg, rgba(255,59,82,0.06) 0%, var(--bg2) 60%); }
.price-card.down { border-left: 3px solid var(--green);  background: linear-gradient(135deg, rgba(0,245,160,0.05) 0%, var(--bg2) 60%); }
.price-card.flat { border-left: 3px solid var(--dim);    }
.price-card.hit-target { border-left: 3px solid var(--gold);   background: linear-gradient(135deg, rgba(255,201,77,0.06) 0%, var(--bg2) 60%); }
.price-card.hit-stop   { border-left: 3px solid var(--red);    background: linear-gradient(135deg, rgba(255,59,82,0.1)  0%, var(--bg2) 60%); }

.section-label {
  font-family: var(--mono) !important;
  font-size: 0.65rem;
  color: var(--muted);
  letter-spacing: 0.14em;
  text-transform: uppercase;
  margin-bottom: 14px;
  display: flex; align-items: center; gap: 10px;
}
.section-label::after {
  content: '';
  flex: 1; height: 1px;
  background: linear-gradient(90deg, var(--line), transparent);
}
.strat-pill {
  display: inline-flex; align-items: center;
  padding: 3px 10px; border-radius: 20px;
  font-size: 0.7rem; font-weight: 600; font-family: var(--mono);
  letter-spacing: 0.06em;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════
# PERSISTENCE LAYER
# ══════════════════════════════════════════════
@st.cache_resource
def _wl(): 
    s = {"codes":[], "groups":{}}
    try:
        if os.path.exists(_WL_FILE):
            d = json.load(open(_WL_FILE))
            s = d if isinstance(d, dict) else {"codes": d, "groups":{}}
    except: pass
    return s

@st.cache_resource
def _notes():
    try: return json.load(open(_NOTES_FILE)) if os.path.exists(_NOTES_FILE) else {}
    except: return {}

@st.cache_resource
def _port():
    try: return json.load(open(_PORT_FILE)) if os.path.exists(_PORT_FILE) else {"trades":[]}
    except: return {"trades":[]}

def _atomic(path, data):
    tmp = path + ".tmp"
    with open(tmp,"w") as f: json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)

import re as _re_v

def _is_valid_code(raw: str) -> bool:
    """Taiwan stock codes are 4-6 digits only (e.g. 2330, 0050, 00878)."""
    bare = str(raw).upper().replace(".TW","").replace(".TWO","").strip()
    # Strip Chinese characters if mixed (e.g. "3019亞光" → "3019")
    bare = _re_v.sub(r'[^\x00-\x7F\d]', '', bare).strip()
    return bool(_re_v.fullmatch(r'\d{4,6}', bare))

def wl_codes():
    raw = _wl()["codes"]
    valid = [c for c in raw if _is_valid_code(c)]
    # Auto-purge: if invalid codes were found, rewrite the file clean
    if len(valid) != len(raw):
        s = _wl()
        s["codes"] = valid
        try: _atomic(_WL_FILE, s)
        except: pass
    return valid
def wl_group(c):         return _wl().get("groups",{}).get(c,"未分組")
def wl_add(c):
    s = _wl(); c = c.upper()
    if c not in s["codes"]: s["codes"].append(c); _atomic(_WL_FILE, s); return True
    return False
def wl_remove(c):
    s = _wl()
    if c in s["codes"]: s["codes"].remove(c); _atomic(_WL_FILE, s); return True
    return False
def wl_set_group(c, g):
    s = _wl(); s.setdefault("groups",{})[c] = g; _atomic(_WL_FILE, s)

def note_get(c):         return _notes().get(c, {})
def note_set(c, d):
    s = _notes(); s[c] = d; _atomic(_NOTES_FILE, s)

def port_trades():       return _port().get("trades", [])
def port_save(trades):
    s = _port(); s["trades"] = trades; _atomic(_PORT_FILE, s)

# ══════════════════════════════════════════════
# CHINESE NAMES TABLE
# ══════════════════════════════════════════════
_NAMES = {
    "2330":"台積電","2317":"鴻海","2454":"聯發科","2881":"富邦金",
    "2308":"台達電","2882":"國泰金","2303":"聯電","2886":"兆豐金",
    "2412":"中華電","2382":"廣達","2891":"中信金","3008":"大立光",
    "2603":"長榮","1301":"台塑","1303":"南亞","2892":"第一金",
    "2002":"中鋼","5880":"合庫金","2207":"和泰車","2395":"研華",
    "2357":"華碩","2887":"台新金","2327":"國巨","2885":"元大金",
    "2884":"玉山金","6415":"矽力-KY","5871":"中租-KY","4938":"臻鼎-KY",
    "2633":"台灣高鐵","3034":"聯詠","3711":"日月光","2301":"光寶科",
    "2912":"統一超","1216":"統一","0050":"元大台灣50","00878":"國泰永續高息",
    "00919":"群益精選高息","2880":"華南金","2883":"開發金","2890":"永豐金",
    "3019":"亞光","2379":"瑞昱","2347":"聯強","2474":"可成",
    "3045":"台灣大","9921":"巨大","9951":"皇田","2542":"興富發",
}
def cn(c): return _NAMES.get(c.upper().replace(".TW","").replace(".TWO",""), c)

# ══════════════════════════════════════════════
# DATA FETCHING
# ══════════════════════════════════════════════
@st.cache_data(ttl=60)
def quote(sym):
    bare = sym.upper().replace(".TW","").replace(".TWO","")
    for s in [bare+".TW", bare+".TWO"]:
        try:
            fi = yf.Ticker(s).fast_info
            px = getattr(fi,"last_price",0) or 0
            if not px: continue
            pv = getattr(fi,"previous_close",px) or px
            return {"px":round(px,2), "prev":round(pv,2),
                    "chg_p":round((px-pv)/pv*100,2) if pv else 0,
                    "hi":getattr(fi,"day_high",0) or 0,
                    "lo":getattr(fi,"day_low",0)  or 0,
                    "vol":getattr(fi,"last_volume",0) or 0, "ok":True}
        except: pass
    return {"ok":False,"px":0,"chg_p":0,"hi":0,"lo":0,"vol":0}

@st.cache_data(ttl=300)
def hist(sym, period="1mo"):
    bare = sym.upper().replace(".TW","").replace(".TWO","")
    for s in [bare+".TW", bare+".TWO"]:
        try:
            df = yf.Ticker(s).history(period=period)
            if not df.empty: return df
        except: pass
    return pd.DataFrame()

@st.cache_data(ttl=60)
def bulk_quotes(codes_tuple):
    return {c: quote(c) for c in codes_tuple}

# ══════════════════════════════════════════════
# TELEGRAM
# ══════════════════════════════════════════════
def tg_send(token, chat_id, text):
    if not token or not chat_id: return False, "未設定"
    try:
        data = _parse.urlencode({"chat_id":str(chat_id).strip(),
                                  "text":_re.sub(r'<[^>]+>','',text)}).encode()
        req = _req.Request(f"https://api.telegram.org/bot{token.strip()}/sendMessage",
                           data=data, method="POST")
        req.add_header("Content-Type","application/x-www-form-urlencoded")
        with _req.urlopen(req, timeout=10) as r:
            return (True,"") if r.status==200 else (False,f"HTTP {r.status}")
    except Exception as e:
        err=str(e)
        if hasattr(e,'read'):
            try: err+=" | "+e.read().decode()[:200]
            except: pass
        return False,err

def alert_sent_today(code, atype):
    try:
        sent = json.load(open(_ALERT_FILE)) if os.path.exists(_ALERT_FILE) else {}
        return sent.get(f"{code}_{atype}") == today_str()
    except: return False

def mark_alert_sent(code, atype):
    try:
        sent = json.load(open(_ALERT_FILE)) if os.path.exists(_ALERT_FILE) else {}
        sent[f"{code}_{atype}"] = today_str()
        _atomic(_ALERT_FILE, sent)
    except: pass

def run_price_alerts(token, chat_id, codes, quotes_d):
    sent = []
    for c in codes:
        n = note_get(c); q = quotes_d.get(c,{})
        px = q.get("px",0)
        if not px: continue
        tgt, stp = n.get("target"), n.get("stop")
        name = cn(c); chg = q.get("chg_p",0)
        def push(atype, msg):
            if not alert_sent_today(c, atype):
                ok,_ = tg_send(token, chat_id, msg)
                if ok: mark_alert_sent(c, atype); sent.append(msg.split('\n')[0])
        if tgt and px >= tgt:
            push("target", f"🎯 目標價達標\n{c} {name}\n現價 {px:.2f}  目標 {tgt}\n{chg:+.2f}%  {now_tw().strftime('%H:%M')}")
        if stp and px <= stp:
            push("stop", f"🛑 停損觸發\n{c} {name}\n現價 {px:.2f}  停損 {stp}\n{chg:+.2f}%  {now_tw().strftime('%H:%M')}")
        if tgt and px < tgt and (tgt-px)/tgt*100 <= 2:
            push("near", f"⚠️ 接近目標\n{c} {name}\n現價 {px:.2f}  距目標 {(tgt-px)/tgt*100:.1f}%")
    return sent

# ══════════════════════════════════════════════
# STRATEGY ENGINE
# ══════════════════════════════════════════════
def evaluate(code, trades, cur_px):
    buys  = [t for t in trades if t.get("交易別","") in ("現買","買進","買")
             and str(t.get("商品","")).startswith(code)]
    sells = [t for t in trades if t.get("交易別","") in ("現賣","賣出","賣")
             and str(t.get("商品","")).startswith(code)]
    held  = sum(t.get("股數",0) for t in buys) - sum(t.get("股數",0) for t in sells)
    if held <= 0: return None
    if not cur_px: return None

    bq = sum(t.get("股數",0) for t in buys)
    avg = sum(t.get("股數",0)*t.get("成交價",0)+t.get("手續費",0) for t in buys)/bq if bq else 0
    mkt = cur_px * held
    cst = avg * held
    unrl  = mkt - cst
    unrl_p = unrl/cst*100 if cst else 0
    stop7  = round(avg*0.93,1)

    prices = sorted([t.get("成交價",0) for t in buys])
    dca = f"{len(buys)} 筆買入  均價 {avg:.1f}  最低 {min(prices):.1f}  最高 {max(prices):.1f}" if prices else ""

    if cur_px <= stop7 and unrl_p < -10:
        act, clr, urg = "🚨 立即停損", "#ff4455", "URGENT"
        reason = f"已跌破 -7% 停損位（{stop7}），虧損 {unrl_p:.1f}%。建議清倉 {held:,} 股，控制損失。"
    elif unrl_p < -15:
        act, clr, urg = "🔴 深度虧損", "#ff4455", "HIGH"
        reason = f"虧損已達 {unrl_p:.1f}%，超過 -15% 警戒線。建議先減碼 50%（賣 {held//2:,} 股）。"
    elif unrl_p < -7:
        act, clr, urg = "🟡 審視停損", "#ff9100", "MED"
        reason = f"虧損 {unrl_p:.1f}%，已有 {len(buys)} 次買入。確認技術面是否底部，否則設定停損執行。"
    elif unrl_p >= 25:
        act, clr, urg = "🟡 分批出場", "#ff9100", "MED"
        reason = f"獲利已達 {unrl_p:.1f}%，建議先賣 {held//2:,} 股鎖定利潤，剩餘設移動停損。"
    elif unrl_p >= 12:
        act, clr, urg = "🟢 續抱觀察", "#00e676", "LOW"
        reason = f"獲利 {unrl_p:.1f}%，趨勢健康。設移動停損 {round(cur_px*0.95,1)}，讓利潤奔跑。"
    elif cur_px <= min(prices)*1.02 and len(buys) < 12:
        act, clr, urg = "🔵 底部加碼？", "#00c8ff", "LOW"
        reason = f"現價 {cur_px:.1f} 接近最低買入成本 {min(prices):.1f}，若基本面未變可評估小量加碼。"
    else:
        act, clr, urg = "⚪ 持有觀察", "#5d7a94", "LOW"
        reason = f"損益 {unrl_p:+.1f}%，無明確操作訊號。等待技術面確認方向再行動。"

    return {"code":code,"name":cn(code),"held":held,"avg":round(avg,2),
            "cur_px":cur_px,"cst":round(cst),"mkt":round(mkt),
            "unrl":round(unrl),"unrl_p":round(unrl_p,2),
            "stop7":stop7,"action":act,"color":clr,"urgency":urg,
            "reason":reason,"dca":dca,
            "first_buy":min(t.get("交易日","") for t in buys),
            "last_buy": max(t.get("交易日","") for t in buys),}

# ══════════════════════════════════════════════
# GOOGLE SHEETS
# ══════════════════════════════════════════════
_GS = "https://sheets.googleapis.com/v4/spreadsheets"

def gs_push(sheet_id, token, trades):
    if not sheet_id or not token: return False, "設定未完成"
    hdr = [["商品","交易日","交易別","股數","成交價","價金","手續費","交易稅","備註"]]
    rows = hdr + [[t.get(k,"") for k in ["商品","交易日","交易別","股數","成交價","價金","手續費","交易稅","備註"]] for t in trades]
    try:
        for url, body, method in [
            (f"{_GS}/{sheet_id}/values/Portfolio!A1:I10000:clear", b"{}", "POST"),
            (f"{_GS}/{sheet_id}/values/Portfolio!A1?valueInputOption=USER_ENTERED",
             json.dumps({"values":rows}).encode(), "PUT"),
        ]:
            r = _req.Request(url, data=body, method=method)
            r.add_header("Authorization", f"Bearer {token.strip()}")
            r.add_header("Content-Type", "application/json")
            with _req.urlopen(r, timeout=10) as resp:
                if resp.status not in (200,): return False, f"HTTP {resp.status}"
        return True, ""
    except Exception as e:
        err=str(e)
        if hasattr(e,'read'):
            try: err+=" | "+e.read().decode()[:300]
            except: pass
        return False, err

# ══════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════
_BG    = PatternFill("solid", fgColor="0D1420")
_HDR   = PatternFill("solid", fgColor="111D2E")
_THIN  = Side(style="thin", color="1A2D44")
_BRD   = Border(left=_THIN,right=_THIN,top=_THIN,bottom=_THIN)
_C     = Alignment(horizontal="center",vertical="center")
_R     = Alignment(horizontal="right", vertical="center")
_WF    = Font(color="CDD9E8")
_MF    = Font(color="5D7A94")
_HF    = Font(color="CDD9E8", bold=True)
_AF    = Font(color="00C8FF", bold=True)
_GF    = Font(color="00E676", bold=True)
_RF    = Font(color="FF4455", bold=True)

def trades_to_excel(trades: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "買賣記錄"
    cols = ["商品","交易日","交易別","股數","成交價","價金","手續費","交易稅","備註"]
    for ci,h in enumerate(cols,1):
        c = ws.cell(1,ci,h); c.font=_HF; c.fill=_HDR; c.alignment=_C; c.border=_BRD
    ws.row_dimensions[1].height = 22
    for ri,t in enumerate(trades,2):
        for ci,k in enumerate(cols,1):
            v = t.get(k,"")
            c = ws.cell(ri,ci,v); c.fill=_BG; c.border=_BRD
            if ci==1: c.font=_RF; c.alignment=_C
            elif ci==3: c.font=Font(color="FF9100",bold=True); c.alignment=_C
            elif ci in(4,7,8): c.font=_MF; c.alignment=_C
            else: c.font=_WF; c.alignment=_R
        ws.row_dimensions[ri].height = 20
    for ci,w in enumerate([14,14,7,6,9,9,8,8,18],1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

PRESET_GROUPS = ["未分組","半導體","金融","電子","傳產","ETF","觀察中"]

# ══════════════════════════════════════════════════════════
# ──  HEADER  ──
# ══════════════════════════════════════════════════════════
st.markdown(f"""
<div style="border-bottom:1px solid #1a2d44;padding-bottom:12px;margin-bottom:18px">
  <div style="display:flex;align-items:baseline;gap:10px;flex-wrap:wrap">
    <span style="font-family:'IBM Plex Mono',monospace;font-size:1.25rem;
      font-weight:700;color:#cdd9e8">WATCHLIST&nbsp;<span style="color:#00c8ff">PRO</span></span>
    <span style="font-size:0.68rem;color:#5d7a94;letter-spacing:0.08em">
      v{APP_VERSION} &nbsp;·&nbsp; {now_tw().strftime("%m/%d %H:%M")}
    </span>
  </div>
  <div style="font-size:0.7rem;color:#3d5470;margin-top:3px;letter-spacing:0.04em">
    台股自選股管理系統
  </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# ──  SIDEBAR  ──
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.72rem;color:#5d7a94;'
                'letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px">'
                '◈ MANAGE WATCHLIST</div>', unsafe_allow_html=True)

    # Add stock
    with st.expander("➕ 新增股票", expanded=True):
        nc = st.text_input("代號", placeholder="2330  /  3019  /  6415.TWO", label_visibility="collapsed")
        ng = st.selectbox("分組", PRESET_GROUPS, label_visibility="collapsed")
        if st.button("新增", type="primary", width='stretch'):
            bare = nc.strip().upper().replace(".TW","").replace(".TWO","")
            if bare:
                if wl_add(bare): wl_set_group(bare, ng); st.success(f"✓ {bare}  {cn(bare)}"); st.rerun()
                else: st.warning(f"{bare} 已在清單")

    # Remove stock
    codes = wl_codes()
    if codes:
        with st.expander("🗑 移除股票"):
            dc = st.selectbox("選擇", codes, format_func=lambda x: f"{x}  {cn(x)}", label_visibility="collapsed")
            if st.button("從清單移除", width='stretch'):
                wl_remove(dc); st.success(f"已移除 {dc}"); st.rerun()

        with st.expander("🗂 調整分組"):
            mc = st.selectbox("股票", codes, format_func=lambda x:f"{x} {cn(x)}", key="mg_c", label_visibility="collapsed")
            mg = st.selectbox("分組", PRESET_GROUPS, key="mg_g", label_visibility="collapsed")
            if st.button("更新", width='stretch'):
                wl_set_group(mc, mg); st.success("✓"); st.rerun()

    st.markdown('<div style="height:1px;background:#1a2d44;margin:16px 0"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.72rem;color:#5d7a94;'
                'letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px">'
                '◈ NOTIFICATIONS</div>', unsafe_allow_html=True)

    # Telegram
    with st.expander("📲 Telegram 推播"):
        _ts = {}
        try: _ts = st.secrets.get("telegram",{})
        except: pass
        tg_tok = st.text_input("Bot Token", value=_ts.get("token",""), type="password", key="tg_tok")
        tg_cid = st.text_input("Chat ID",   value=_ts.get("chat_id",""), key="tg_cid")
        if tg_tok:
            st.markdown(f'<a href="https://api.telegram.org/bot{tg_tok}/getUpdates" '
                       f'target="_blank" style="font-size:0.72rem;color:#00c8ff">'
                       f'↗ getUpdates（查 Chat ID）</a>', unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        if c1.button("測試", width='stretch'):
            ok,e = tg_send(tg_tok, tg_cid, f"✅ Watchlist Pro 推播測試\n{now_tw().strftime('%H:%M')}")
            (st.success("✓") if ok else st.error(e))
        near_alert = c2.checkbox("接近提醒", value=True)

    st.markdown('<div style="height:1px;background:#1a2d44;margin:16px 0"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.72rem;color:#5d7a94;'
                'letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px">'
                '◈ DATA</div>', unsafe_allow_html=True)

    # Import/Export
    with st.expander("📁 匯入 / 匯出"):
        if codes:
            df_exp = pd.DataFrame({"代號":codes,"名稱":[cn(c) for c in codes],"分組":[wl_group(c) for c in codes]})
            buf2 = io.BytesIO()
            wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title="自選股"
            ws2.append(["代號","名稱","分組"])
            for _,r in df_exp.iterrows(): ws2.append([r["代號"],r["名稱"],r["分組"]])
            wb2.save(buf2); buf2.seek(0)
            st.download_button("↓ 匯出自選股 Excel", data=buf2,
                file_name=f"watchlist_{today_str()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch')
        uf = st.file_uploader("↑ 匯入 Excel", type=["xlsx"], label_visibility="collapsed")
        if uf:
            try:
                df_i = pd.read_excel(uf, sheet_name=0)
                added=0
                for _,r in df_i.iterrows():
                    c=str(r.get("代號",r.get("code",""))).strip().upper()
                    g=str(r.get("分組","未分組")).strip()
                    if c and wl_add(c): wl_set_group(c,g); added+=1
                st.success(f"匯入 {added} 支"); st.rerun()
            except Exception as e: st.error(str(e))

    # Auto refresh
    with st.expander("⏱ 自動更新"):
        auto = st.toggle("啟用", value=False)
        if auto:
            ivl = st.select_slider("間隔", ["30秒","1分","5分"], value="1分")
            secs = {"30秒":30,"1分":60,"5分":300}[ivl]
            st.markdown(f'<meta http-equiv="refresh" content="{secs}">', unsafe_allow_html=True)
            st.caption(f"每 {ivl} 自動刷新")
    if st.button("↻ 立即刷新報價", width='stretch'):
        st.cache_data.clear(); st.rerun()

# ══════════════════════════════════════════════════════════
# ──  EMERGENCY RESET — always visible, before any st.stop()  ──
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div style="height:1px;background:#1a2d44;margin:16px 0"></div>',
                unsafe_allow_html=True)
    with st.expander("⚠️ 緊急重置", expanded=False):
        st.caption("若 App 卡住或資料異常，點此清除 /tmp 資料並重載")
        if st.button("🗑 清除所有暫存資料", width='stretch', key="emergency_reset"):
            for f in [_WL_FILE, _NOTES_FILE, _PORT_FILE, _ALERT_FILE]:
                try: os.remove(f)
                except: pass
            st.cache_data.clear()
            st.cache_resource.clear()
            st.success("已清除，重新載入中…")
            st.rerun()
        if st.button("🔄 只清除自選股清單", width='stretch', key="reset_wl"):
            try: os.remove(_WL_FILE)
            except: pass
            st.cache_resource.clear()
            st.rerun()

# ══════════════════════════════════════════════════════════
# ──  GUARD: no stocks  ──
# ══════════════════════════════════════════════════════════
codes = wl_codes()
if not codes:
    st.markdown("""
    <div style="text-align:center;padding:40px 0 20px 0">
      <div style="font-family:'IBM Plex Mono',monospace;font-size:0.9rem;
        color:#5d7a94;letter-spacing:0.12em">WATCHLIST EMPTY</div>
      <div style="color:#3d5470;font-size:0.82rem;margin-top:6px">新增股票代號開始使用</div>
    </div>""", unsafe_allow_html=True)

    # Inline add form — mobile-first stacked layout
    with st.container():
        in_code  = st.text_input("股票代號", placeholder="輸入代號  例：2330 / 3019 / 0050",
                                  label_visibility="collapsed", key="inline_code")
        in_group = st.selectbox("分組", PRESET_GROUPS,
                                 label_visibility="collapsed", key="inline_group")
        if st.button("➕  新增到自選股", type="primary", width='stretch', key="inline_add"):
            bare = in_code.strip().upper().replace(".TW","").replace(".TWO","")
            if bare and _is_valid_code(bare):
                wl_add(bare); wl_set_group(bare, in_group)
                st.success(f"✓ 已新增 {bare}  {cn(bare)}"); st.rerun()
            elif bare:
                st.error(f"「{bare}」格式不正確（台股代號為 4-6 位數字）")
            else:
                st.warning("請輸入股票代號")

    st.markdown("""
    <div style="text-align:center;margin-top:16px;color:#3d5470;font-size:0.78rem">
      常用代號：2330 台積電 ／ 2317 鴻海 ／ 0050 元大台灣50 ／ 3019 亞光
    </div>""", unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════
# ──  FETCH ALL QUOTES  ──
# ══════════════════════════════════════════════════════════
with st.spinner(""):
    Q = bulk_quotes(tuple(codes))

# Auto price alerts
_tok = st.session_state.get("tg_tok","")
_cid = st.session_state.get("tg_cid","")
if _tok and _cid:
    sent_alerts = run_price_alerts(_tok, _cid, codes, Q)
    for a in sent_alerts: st.toast(a, icon="📲")

# ══════════════════════════════════════════════════════════
# ──  TABS  ──
# ══════════════════════════════════════════════════════════
tab_ov, tab_card, tab_mon, tab_note, tab_alert, tab_port = st.tabs([
    "OVERVIEW", "CARDS", "MONITOR", "NOTES", "ALERTS", "PORTFOLIO"
])

# ────────────────────────────────────────────────────────
# TAB 1 — OVERVIEW TABLE
# ────────────────────────────────────────────────────────
with tab_ov:
    grp_all = sorted(set(wl_group(c) for c in codes))
    fg = st.selectbox("分組篩選", ["ALL"] + grp_all, label_visibility="collapsed", key="ov_grp")

    rows = []
    for c in codes:
        if fg != "ALL" and wl_group(c) != fg: continue
        q = Q.get(c,{}); n = note_get(c)
        px = q.get("px",0); chg = q.get("chg_p",0)
        tgt = n.get("target"); stp = n.get("stop")
        up = round((tgt/px-1)*100,1) if tgt and px else None
        rows.append({
            "代號":    c,
            "名稱":    cn(c),
            "分組":    wl_group(c),
            "現價":    px or None,
            "漲跌%":   chg if q.get("ok") else None,
            "目標價":  float(tgt) if tgt else None,
            "停損價":  float(stp) if stp else None,
            "上漲空間%": up,
            "狀態":    ("🎯" if tgt and px and px>=tgt else
                        "🛑" if stp and px and px<=stp else
                        "⚠️" if tgt and px and up and 0<up<=5 else "─"),
        })

    if rows:
        df_ov = pd.DataFrame(rows)
        # Summary
        valid = [r for r in rows if r["漲跌%"] is not None]
        m1,m2 = st.columns(2); m3,m4 = st.columns(2)
        m1.metric("總持倉", len(codes))
        m2.metric("警報", sum(1 for r in rows if r["狀態"]!="─"))
        m3.metric("🔴 上漲", sum(1 for r in valid if r["漲跌%"]>0))
        m4.metric("🟢 下跌", sum(1 for r in valid if r["漲跌%"]<0))
        st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)
        st.dataframe(df_ov, width='stretch', hide_index=True,
            column_config={
                "漲跌%": st.column_config.NumberColumn(format="%+.2f%%"),
                "上漲空間%": st.column_config.NumberColumn(format="%+.1f%%"),
                "現價": st.column_config.NumberColumn(format="%.2f"),
            })

# ────────────────────────────────────────────────────────
# TAB 2 — CARDS
# ────────────────────────────────────────────────────────
with tab_card:
    col_sort = st.radio("排序", ["預設","漲幅↓","跌幅↓","分組"], horizontal=True, label_visibility="collapsed")
    sc = list(codes)
    if col_sort == "漲幅↓": sc.sort(key=lambda x: -Q.get(x,{}).get("chg_p",0))
    elif col_sort == "跌幅↓": sc.sort(key=lambda x: Q.get(x,{}).get("chg_p",0))
    elif col_sort == "分組": sc.sort(key=lambda x: wl_group(x))

    # 2 columns on mobile, 3 on desktop
    n_cols = 2
    for i in range(0, len(sc), n_cols):
        chunk = sc[i:i+n_cols]
        cols = st.columns(n_cols)
        for ci, c in enumerate(chunk):
            q = Q.get(c,{}); n = note_get(c)
            px = q.get("px",0); chg = q.get("chg_p",0)
            tgt = n.get("target"); stp = n.get("stop")
            hi = q.get("hi",0);   lo  = q.get("lo",0)

            if not q.get("ok"):    clr, card_cls = "var(--muted)", "flat"
            elif chg > 0:          clr, card_cls = "var(--red)",   "up"
            elif chg < 0:          clr, card_cls = "var(--green)", "down"
            else:                  clr, card_cls = "var(--muted)", "flat"
            if tgt and px and px>=tgt: card_cls = "hit-target"; clr = "var(--gold)"
            if stp and px and px<=stp: card_cls = "hit-stop";   clr = "var(--red)"

            chg_s  = f"+{chg:.2f}%" if chg>0 else f"{chg:.2f}%"
            px_s   = f"{px:.2f}"   if px  else "─"
            hi_s   = f"{hi:.2f}"   if hi  else "─"
            lo_s   = f"{lo:.2f}"   if lo  else "─"
            tgt_s  = f"🎯 {tgt}"  if tgt else ""
            stp_s  = f"🛑 {stp}"  if stp else ""

            with cols[ci]:
                st.markdown(f"""
<div class="wl-card {card_cls}">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px">
    <span class="ticker-mono" style="color:#cdd9e8;font-size:0.95rem">{c}</span>
    <span style="font-size:0.7rem;color:#5d7a94;background:#111d2e;
      padding:2px 8px;border-radius:4px;letter-spacing:0.04em">{wl_group(c)}</span>
  </div>
  <div style="font-size:0.78rem;color:#5d7a94;margin-bottom:8px">{cn(c)}</div>
  <div style="display:flex;align-items:baseline;gap:8px;margin-bottom:6px">
    <span style="font-family:'IBM Plex Mono';font-size:1.3rem;font-weight:700;color:{clr}">{px_s}</span>
    <span style="font-size:0.8rem;color:{clr}">{chg_s if q.get('ok') else '─'}</span>
  </div>
  <div style="font-size:0.72rem;color:#3d5470;display:flex;gap:12px;flex-wrap:wrap">
    <span>H {hi_s}</span><span>L {lo_s}</span>
    {f'<span style="color:#ffd600">{tgt_s}</span>' if tgt_s else ''}
    {f'<span style="color:#ff4455">{stp_s}</span>' if stp_s else ''}
  </div>
  {f'<div style="font-size:0.7rem;color:#3d5470;margin-top:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{n.get("note","")[:40]}</div>' if n.get("note") else ''}
</div>""", unsafe_allow_html=True)

# ────────────────────────────────────────────────────────
# TAB 3 — MONITOR (daily chart)
# ────────────────────────────────────────────────────────
with tab_mon:
    valid_q = [c for c in codes if Q.get(c,{}).get("ok")]
    if valid_q:
        sorted_q = sorted(valid_q, key=lambda c: -Q[c]["chg_p"])
        top3 = sorted_q[:3]; bot3 = sorted_q[-3:][::-1]
        g1,g2 = st.columns(2)
        for col,title,lst,clr in [(g1,"🔴 漲幅前三",top3,"#ff4455"),(g2,"🟢 跌幅前三",bot3,"#00e676")]:
            with col:
                st.markdown(f'<div style="font-size:0.72rem;color:#5d7a94;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:8px">{title}</div>', unsafe_allow_html=True)
                for c in lst:
                    q = Q[c]; chg=q["chg_p"]; px=q["px"]
                    c_icon = "▲" if chg>0 else "▼" if chg<0 else "─"
                    st.markdown(f"""
<div style="background:var(--bg2);border-left:3px solid {clr};
  padding:10px 14px;border-radius:0 var(--r8) var(--r8) 0;margin-bottom:8px;
  display:flex;justify-content:space-between;align-items:center;
  border-top:1px solid var(--line2);border-right:1px solid var(--line2);
  border-bottom:1px solid var(--line2)">
  <span>
    <span style="font-family:var(--mono);font-weight:700;color:var(--text);
      font-size:0.88rem">{c}</span>
    <span style="font-size:0.75rem;color:var(--muted);margin-left:8px;
      font-weight:500">{cn(c)}</span>
  </span>
  <span style="font-family:var(--mono);font-weight:700;color:{clr};font-size:0.88rem;
    display:flex;gap:10px;align-items:center">
    <span>{chg:+.2f}%</span>
    <span style="color:var(--text)">{px:.2f}</span>
  </span>
</div>""", unsafe_allow_html=True)

    st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
    chart_c = st.selectbox("選股", codes, format_func=lambda x:f"{x}  {cn(x)}", label_visibility="collapsed", key="mon_c")
    prd = st.radio("區間", ["1mo","3mo","6mo","1y"], horizontal=True, label_visibility="collapsed", key="mon_prd")
    with st.spinner(""):
        df_h = hist(chart_c, prd)
    if not df_h.empty:
        n_mon = note_get(chart_c)
        vol_avg = df_h["Volume"].mean()
        fig = make_subplots(rows=2,cols=1,shared_xaxes=True,row_heights=[0.72,0.28],
                            vertical_spacing=0.04,
                            specs=[[{"secondary_y":False}],[{"secondary_y":False}]])
        fig.add_trace(go.Candlestick(
            x=df_h.index,open=df_h["Open"],high=df_h["High"],low=df_h["Low"],close=df_h["Close"],
            increasing_fillcolor="#ff4455",increasing_line_color="#ff4455",
            decreasing_fillcolor="#00e676",decreasing_line_color="#00e676",
            showlegend=False),row=1,col=1)
        if n_mon.get("target"):
            fig.add_hline(y=n_mon["target"],line_dash="dot",line_color="#ffd600",line_width=1.5,
                annotation_text=f"Target {n_mon['target']}",annotation_font_color="#ffd600",
                annotation_position="right",row=1,col=1)
        if n_mon.get("stop"):
            fig.add_hline(y=n_mon["stop"],line_dash="dot",line_color="#ff4455",line_width=1.5,
                annotation_text=f"Stop {n_mon['stop']}",annotation_font_color="#ff4455",
                annotation_position="right",row=1,col=1)
        v_clr = ["#ff4455" if df_h["Close"].iloc[i]>=df_h["Open"].iloc[i] else "#00e676"
                 for i in range(len(df_h))]
        fig.add_trace(go.Bar(x=df_h.index,y=df_h["Volume"],marker_color=v_clr,showlegend=False),row=2,col=1)
        spk = df_h[df_h["Volume"]>vol_avg*2]
        if not spk.empty:
            fig.add_trace(go.Scatter(x=spk.index,y=spk["Volume"],mode="markers",
                marker=dict(color="#ffd600",size=8,symbol="diamond"),name="量異常"),row=2,col=1)
        fig.update_layout(template="plotly_dark",height=400,paper_bgcolor="#080c14",
            plot_bgcolor="#0d1420",margin=dict(l=50,r=80,t=8,b=8),
            font=dict(family="IBM Plex Mono",size=10,color="#5d7a94"),
            hovermode="x unified",xaxis_rangeslider_visible=False,
            legend=dict(orientation="h",y=1.02))
        fig.update_yaxes(gridcolor="#111d2e",showgrid=True)
        fig.update_xaxes(gridcolor="#111d2e",showgrid=False)
        st.plotly_chart(fig, width='stretch')
        last_vol = df_h["Volume"].iloc[-1]
        if last_vol > vol_avg*2:
            st.warning(f"⚠️ 今日成交量 {last_vol:,.0f}，為近期均量 {last_vol/vol_avg:.1f} 倍 — 異常放量")
    else:
        st.info("無歷史資料")

# ────────────────────────────────────────────────────────
# TAB 4 — NOTES
# ────────────────────────────────────────────────────────
with tab_note:
    nc_sel = st.selectbox("選股", codes, format_func=lambda x:f"{x}  {cn(x)}", label_visibility="collapsed", key="note_sel")
    ex = note_get(nc_sel)
    with st.form("note_form"):
        note_t = st.text_area("備忘 / 進場理由", value=ex.get("note",""), height=90,
                               placeholder="記錄關注理由、進場邏輯、觀察重點…")
        nc1,nc2 = st.columns(2)
        entry_px = nc1.number_input("進場均價", min_value=0.0, value=float(ex.get("entry",0) or 0), step=0.5)
        w_date   = nc2.text_input("關注日期",  value=ex.get("watch_date", today_str()))
        tags = st.multiselect("標籤",
            ["技術突破","籌碼轉強","法人買超","業績成長","低估值","高殖利率","週期底部","題材","其他"],
            default=ex.get("tags",[]))
        if st.form_submit_button("💾 儲存", type="primary"):
            note_set(nc_sel,{**ex,"note":note_t,"entry":entry_px,"watch_date":w_date,
                              "tags":tags,"updated":now_tw().strftime("%Y-%m-%d %H:%M")})
            st.success("✓ 已儲存")
    st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.72rem;color:#5d7a94;letter-spacing:0.1em;'
                'text-transform:uppercase;margin-bottom:10px">ALL NOTES</div>', unsafe_allow_html=True)
    nrows = [{"代號":c,"名稱":cn(c),"備忘":note_get(c).get("note","")[:40],
               "標籤":"、".join(note_get(c).get("tags",[])),
               "均價":note_get(c).get("entry",""),"關注日":note_get(c).get("watch_date","")}
             for c in codes if note_get(c).get("note") or note_get(c).get("tags")]
    if nrows: st.dataframe(pd.DataFrame(nrows), width='stretch', hide_index=True)
    else: st.info("尚無備忘記錄")

# ────────────────────────────────────────────────────────
# TAB 5 — ALERTS
# ────────────────────────────────────────────────────────
with tab_alert:
    _tok2 = st.session_state.get("tg_tok","")
    _cid2 = st.session_state.get("tg_cid","")
    if _tok2 and _cid2:
        st.success("📲 Telegram 推播已啟用 — 報價刷新時自動檢查")
    else:
        st.info("💡 在左側側欄設定 Telegram Token 和 Chat ID 以啟用自動推播")

    with st.expander("✏️ 設定目標價 / 停損價", expanded=True):
        ac = st.selectbox("選股", codes, format_func=lambda x:f"{x}  {cn(x)}", label_visibility="collapsed", key="al_c")
        ea = note_get(ac); qa = Q.get(ac,{})
        cpx = qa.get("px",0)
        acol1,acol2,acol3 = st.columns(3)
        tgt_px = acol1.number_input("🎯 目標價", min_value=0.0, step=0.5,
                                     value=float(ea.get("target",0) or 0))
        stp_px = acol2.number_input("🛑 停損價", min_value=0.0, step=0.5,
                                     value=float(ea.get("stop",0) or 0))
        if cpx and tgt_px:
            up2 = (tgt_px/cpx-1)*100
            acol3.metric("上漲空間", f"{up2:+.1f}%", "設定中")
        if st.button("💾 儲存設定", type="primary", key="save_al"):
            note_set(ac,{**ea,"target":tgt_px or None,"stop":stp_px or None})
            st.success(f"✓ {ac} 目標/停損已設定"); st.rerun()

    al_rows = []
    for c in codes:
        n=note_get(c); q=Q.get(c,{}); px=q.get("px",0)
        t=n.get("target"); s=n.get("stop")
        if not(t or s): continue
        up3 = round((t/px-1)*100,1) if t and px else None
        st_list = []
        if t and px and px>=t:               st_list.append("🎯 已達目標")
        if s and px and px<=s:               st_list.append("🛑 跌破停損")
        if t and px and up3 and 0<up3<=3:    st_list.append("⚠️ 極近目標")
        if s and px and px and s and (px-s)/px*100<=3: st_list.append("⚠️ 接近停損")
        if not st_list: st_list.append("✅ 正常")
        al_rows.append({"代號":c,"名稱":cn(c),"現價":round(px,2) if px else None,
                         "目標":float(t) if t else None,"停損":float(s) if s else None,
                         "上漲空間":f"{up3:+.1f}%" if up3 is not None else None,
                         "狀態":" ".join(st_list)})
    if al_rows:
        st.dataframe(pd.DataFrame(al_rows), width='stretch', hide_index=True,
                     column_config={"現價":st.column_config.NumberColumn(format="%.2f")})
        triggered = [r for r in al_rows if "已達" in r["狀態"] or "跌破" in r["狀態"]]
        if triggered:
            st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
            for r in triggered:
                clr = "#ffd600" if "目標" in r["狀態"] else "#ff4455"
                st.markdown(f'<div style="border:1px solid {clr};border-radius:8px;padding:10px 14px;'
                            f'margin-bottom:8px;background:rgba({",".join(str(int(clr.lstrip("#")[i:i+2],16)) for i in (0,2,4))},0.08)">'
                            f'<b style="color:{clr}">{r["狀態"]}</b>&emsp;'
                            f'<span style="color:#cdd9e8">{r["代號"]} {r["名稱"]}</span>&emsp;'
                            f'<span style="color:#5d7a94;font-family:IBM Plex Mono">{r["現價"]}</span>'
                            f'</div>', unsafe_allow_html=True)
    else:
        st.info("尚未對任何股票設定目標價或停損價")

# ────────────────────────────────────────────────────────
# TAB 6 — PORTFOLIO
# ────────────────────────────────────────────────────────
with tab_port:
    trades_all = port_trades()

    # ── IMPORT ──
    with st.expander("📥 匯入買賣記錄", expanded=not trades_all):
        st.markdown("""<div style="font-size:0.8rem;color:#5d7a94;margin-bottom:10px">
        支援格式：<span style="font-family:IBM Plex Mono;color:#00c8ff">
        商品 | 交易日 | 交易別 | 股數 | 成交價 | 價金 | 手續費 | 交易稅</span><br>
        從券商 App 匯出 Excel，或直接上傳截圖由 Claude 轉換。</div>""", unsafe_allow_html=True)
        imp_f = st.file_uploader("選擇 Excel", type=["xlsx","xls"], label_visibility="collapsed")
        if imp_f:
            try:
                df_im = None
                for hr in [0,1,2]:
                    try:
                        imp_f.seek(0)
                        df_t = pd.read_excel(imp_f, header=hr, sheet_name=0)
                        if any(k in " ".join(str(c) for c in df_t.columns) for k in ["商品","交易日","代號"]):
                            df_im = df_t; break
                    except: pass
                if df_im is None: imp_f.seek(0); df_im = pd.read_excel(imp_f, sheet_name=0)
                col_map = {"商品":["商品","股票","代號"],"交易日":["交易日","日期"],"交易別":["交易別","買賣"],
                           "股數":["股數","數量"],"成交價":["成交價","價格"],"價金":["價金","金額"],
                           "手續費":["手續費","費用"],"交易稅":["交易稅","稅"]}
                rc = {}
                for tgt,cands in col_map.items():
                    for cand in cands:
                        m=[c for c in df_im.columns if cand in str(c)]
                        if m: rc[tgt]=m[0]; break
                if "商品" not in rc:
                    st.error("找不到「商品」欄位，請確認 Excel 格式"); 
                else:
                    df_m = pd.DataFrame()
                    for tgt,src in rc.items(): df_m[tgt]=df_im[src]
                    for col in ["商品","交易日","交易別","股數","成交價","價金","手續費","交易稅","備註"]:
                        if col not in df_m: df_m[col] = 0 if col in ["股數","成交價","價金","手續費","交易稅"] else ""
                    for nc in ["股數","成交價","價金","手續費","交易稅"]:
                        df_m[nc]=pd.to_numeric(df_m[nc],errors="coerce").fillna(0)
                    st.dataframe(df_m.head(8), width='stretch', hide_index=True)
                    if st.button("✅ 確認匯入", type="primary"):
                        port_save(df_m.to_dict("records")); st.success(f"已匯入 {len(df_m)} 筆"); st.rerun()
            except Exception as e: st.error(str(e))

        st.divider()
        st.markdown('<div style="font-size:0.75rem;color:#5d7a94;margin-bottom:8px">手動新增一筆</div>',unsafe_allow_html=True)
        with st.form("add_trade"):
            r1c1,r1c2,r1c3 = st.columns(3)
            t_c=r1c1.text_input("商品",placeholder="3019亞光")
            t_d=r1c2.date_input("交易日",value=date.today())
            t_t=r1c3.selectbox("交易別",["現買","現賣"])
            _c1,_c2=st.columns(2); _c3,_c4=st.columns(2)
            t_q=_c1.number_input("股數",min_value=1,value=50,step=1)
            t_p=_c2.number_input("成交價",min_value=0.1,value=100.0,step=0.5)
            auto_fee=int(t_q*t_p*0.001425*0.6)
            auto_tax=int(t_q*t_p*0.003) if t_t=="現賣" else 0
            t_f=_c3.number_input("手續費",min_value=0,value=auto_fee)
            t_x=_c4.number_input("交易稅",min_value=0,value=auto_tax)
            if st.form_submit_button("新增",type="primary"):
                ex_t=port_trades()
                ex_t.append({"商品":t_c,"交易日":str(t_d),"交易別":t_t,"股數":int(t_q),
                              "成交價":float(t_p),"價金":int(t_q*t_p),"手續費":int(t_f),"交易稅":int(t_x),"備註":""})
                port_save(ex_t); st.success("✓"); st.rerun()

    if not trades_all:
        st.info("請先匯入或手動新增買賣記錄"); st.stop()

    # ── ANALYSIS & STRATEGY ──
    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.72rem;color:#5d7a94;letter-spacing:0.1em;'
                'text-transform:uppercase;margin-bottom:14px">POSITION ANALYSIS</div>', unsafe_allow_html=True)

    def _valid_code(raw):
        """Only accept codes that are purely numeric (Taiwan stock codes are digits only)."""
        bare = str(raw).upper().replace(".TW","").replace(".TWO","").replace("亞光","").replace("台積電","").strip()
        # Extract leading digits
        import re as _rev
        m = _rev.match(r'^(\d{4,6})', bare)
        return m.group(1) if m else None

    all_codes_p = sorted(set(
        c for c in (
            _valid_code(str(t.get("商品","")))
            for t in trades_all if t.get("商品")
        ) if c
    ))
    evals = []
    for code in all_codes_p:
        q2 = quote(code); px2 = q2.get("px",0)
        if not px2: q2=quote(code+".TW"); px2=q2.get("px",0)
        ev = evaluate(code, trades_all, px2)
        if ev: evals.append(ev)

    if evals:
        tc,tm,tu,tp = sum(e["cst"] for e in evals),sum(e["mkt"] for e in evals),\
                      sum(e["unrl"] for e in evals),0
        tp = tu/tc*100 if tc else 0
        pm1,pm2 = st.columns(2); pm3,pm4 = st.columns(2)
        pm1.metric("持倉支數",  len(evals))
        pm2.metric("總投入",    f"{tc/1e4:.1f}萬")
        pm3.metric("目前市值",  f"{tm/1e4:.1f}萬")
        pm4.metric("損益",      f"{tu/1e4:+.2f}萬", f"{tp:+.1f}%")

        st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)

        # Summary table
        tbl = pd.DataFrame([{"代號":e["code"],"名稱":e["name"],"持股":e["held"],
            "均成本":e["avg"],"現價":e["cur_px"],"損益%":e["unrl_p"],
            "損益(元)":e["unrl"],"建議":e["action"],"緊急":e["urgency"]} for e in evals])
        st.dataframe(tbl,width='stretch',hide_index=True,
            column_config={"損益%":st.column_config.NumberColumn(format="%+.2f%%"),
                           "損益(元)":st.column_config.NumberColumn(format="$%+,.0f")})

        st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:0.72rem;color:#5d7a94;letter-spacing:0.1em;'
                    'text-transform:uppercase;margin-bottom:14px">STRATEGY RECOMMENDATIONS</div>',
                    unsafe_allow_html=True)

        urg_ord = {"URGENT":0,"HIGH":1,"MED":2,"LOW":3}
        for ev in sorted(evals, key=lambda e: urg_ord.get(e["urgency"],4)):
            pnl_c = "#00e676" if ev["unrl_p"]>=0 else "#ff4455"
            urg_label = {"URGENT":"🔴 立即處理","HIGH":"🟠 本週處理","MED":"🟡 注意觀察","LOW":"⚪ 持續觀察"}.get(ev["urgency"],"")
            clr_hex = ev['color'] if ev['color'].startswith('#') else '#00c8ff'
            r,g,b = int(clr_hex[1:3],16),int(clr_hex[3:5],16),int(clr_hex[5:7],16)
            rgb = f"{r},{g},{b}"
            st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba({rgb},0.06) 0%,var(--bg2) 50%);
  border:1px solid rgba({rgb},0.2);border-left:3px solid {ev['color']};
  border-radius:var(--r12);padding:16px 18px;margin-bottom:14px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;margin-bottom:10px">
    <div>
      <span style="font-family:var(--mono);font-size:0.95rem;font-weight:700;color:var(--text)">{ev['code']}</span>
      <span style="font-size:0.8rem;color:var(--muted);margin-left:8px">{ev['name']}</span>
    </div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
      <span style="font-size:0.68rem;font-weight:600;color:{ev['color']};font-family:var(--mono);
        background:rgba({rgb},0.12);border:1px solid rgba({rgb},0.3);
        padding:2px 10px;border-radius:12px">{urg_label}</span>
      <span style="font-family:var(--mono);font-weight:700;color:{ev['color']};font-size:0.9rem">{ev['action']}</span>
    </div>
  </div>
  <div style="font-size:0.82rem;color:var(--muted);line-height:1.65;
    padding:10px 12px;background:var(--bg1);border-radius:var(--r8);
    margin-bottom:12px;border-left:2px solid rgba({rgb},0.35)">{ev['reason']}</div>
  <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:6px;
    font-family:var(--mono);font-size:0.7rem">
    {"".join([
      f'<div style="background:var(--bg1);padding:6px 8px;border-radius:var(--r4);border:1px solid var(--line2);text-align:center"><div style="color:var(--muted);font-size:0.58rem;text-transform:uppercase">{label}</div><div style="color:{color};font-weight:700;margin-top:3px">{val}</div></div>'
      for label,val,color in [
        ("持股",f"{ev['held']:,}","var(--text)"),
        ("均成本",f"{ev['avg']:.2f}","var(--text)"),
        ("現價",f"{ev['cur_px']:.2f}","var(--text)"),
        ("損益",f"{ev['unrl_p']:+.1f}%",pnl_c),
        ("停損",f"{ev['stop7']}","var(--red)"),
      ]
    ])}
  </div>
  <div style="font-size:0.65rem;color:var(--dim);margin-top:8px;padding-top:8px;
    border-top:1px solid var(--line2)">{ev['dca']}  ·  首買 {ev['first_buy']}</div>
</div>""", unsafe_allow_html=True)

    # ── GS SYNC ──
    st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.72rem;color:#5d7a94;letter-spacing:0.1em;'
                'text-transform:uppercase;margin-bottom:14px">SYNC & EXPORT</div>', unsafe_allow_html=True)
    pc1,pc2 = st.columns(2)
    _gs_s = {}
    try: _gs_s = st.secrets.get("gsheets",{})
    except: pass
    gs_id = pc1.text_input("Google Sheet ID", value=_gs_s.get("sheet_id",""), key="gs_id",
        help="URL中 /d/ 後面那段\nhttps://docs.google.com/spreadsheets/d/[THIS_PART]/edit")
    gs_tok= pc2.text_input("OAuth Bearer Token", value=_gs_s.get("token",""),
                            type="password", key="gs_tok",
                            help="從 developers.google.com/oauthplayground 取得（ya29.xxx）")

    with st.expander("🔐 安全設定教學"):
        st.markdown("""
**最安全 → 存入 Streamlit Cloud Secrets（不進 GitHub）**

Streamlit Cloud → 你的 App → `⋮` → Settings → Secrets：
```toml
[telegram]
token   = "你的Bot Token"
chat_id = "你的Chat ID"

[gsheets]
sheet_id = "從Sheet URL複製"
token    = "ya29.OAuth Token"
```

**取得 Sheet ID**
```
https://docs.google.com/spreadsheets/d/【ID在這裡】/edit
```

**取得 OAuth Token（有效1小時）**
1. 開 developers.google.com/oauthplayground
2. 選 Google Sheets API v4 → spreadsheets
3. Authorize → Exchange → 複製 Access token
""")

    gc1,gc2,gc3 = st.columns(3)
    if gc1.button("⬆️ 同步到 Google Sheets", type="primary", width='stretch'):
        with st.spinner("上傳中…"):
            ok,err = gs_push(gs_id, gs_tok, trades_all)
        st.success("✅ 已同步到 Portfolio 工作表") if ok else st.error(f"❌ {err}")

    if gc2.button("📲 推播策略到 Telegram", width='stretch'):
        if not _tok or not _cid: st.warning("請先設定 Telegram")
        else:
            n_sent=0
            for ev in (evals if evals else []):
                msg=(f"{ev['action']}  {ev['code']} {ev['name']}\n"
                     f"現價 {ev['cur_px']:.2f}  均成本 {ev['avg']:.2f}\n"
                     f"損益 {ev['unrl_p']:+.1f}% ({ev['unrl']:+,.0f}元)\n"
                     f"停損位 {ev['stop7']}\n{ev['reason'][:80]}\n"
                     f"{now_tw().strftime('%m/%d %H:%M')}")
                ok2,_=tg_send(_tok,_cid,msg)
                if ok2: n_sent+=1
            st.success(f"✅ 已推播 {n_sent} 支")

    with gc3:
        st.download_button("↓ 下載 Excel",
            data=trades_to_excel(trades_all),
            file_name=f"portfolio_{today_str()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch')

    if st.button("🗑 清除所有交易記錄", width='stretch', key="clear_trades"):
        port_save([]); st.rerun()

# ══════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════
st.markdown(f"""
<div style="border-top:1px solid #1a2d44;margin-top:24px;padding-top:12px;
  display:flex;justify-content:space-between;align-items:center">
  <div style="font-family:IBM Plex Mono;font-size:0.68rem;color:#3d5470">
    WATCHLIST PRO v{APP_VERSION}  ·  {len(codes)} STOCKS  ·  {now_tw().strftime("%H:%M:%S")} CST
  </div>
  <div style="font-family:IBM Plex Mono;font-size:0.68rem;color:#3d5470">
    SENTINEL PRO COMPANION
  </div>
</div>
""", unsafe_allow_html=True)
